from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.http import HttpResponse

import openpyxl
from openpyxl.utils import get_column_letter

from .models import Categoria, Producto
from .forms import ProductoForm, CategoriaForm
from django.contrib.auth.decorators import login_required, permission_required
# ----------------------------------------
# VISTAS CRUD
# ----------------------------------------

# --- CRUD de Categorías ---
@login_required
@permission_required('catalogo.view_categoria', raise_exception=True)
def categoria_list(request):
    nombre_filtro = request.GET.get('nombre', '')
    categorias = Categoria.objects.all()
    if nombre_filtro:
        categorias = categorias.filter(nombre__icontains=nombre_filtro)

    # PAGINACIÓN
    from django.core.paginator import Paginator
    paginator = Paginator(categorias, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'catalogo/categoria_list.html', {'page_obj': page_obj, 'nombre_filtro': nombre_filtro})

@login_required
@permission_required('catalogo.add_categoria', raise_exception=True)
def categoria_create(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('catalogo:categoria_list')
    else:
        form = CategoriaForm()
    return render(request, 'catalogo/categoria_form.html', {'form': form})

@login_required
@permission_required('catalogo.change_categoria', raise_exception=True)
def categoria_update(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('catalogo:categoria_list')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'catalogo/categoria_form.html', {'form': form})

@login_required
@permission_required('catalogo.delete_categoria', raise_exception=True)
def categoria_delete(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        nombre_cat = categoria.nombre
        categoria.delete()
        messages.success(request, f'Categoría {nombre_cat} eliminada.')
        return redirect('catalogo:categoria_list')
    return render(request, 'catalogo/categoria_confirm_delete.html', {'object': categoria})


# --- CRUD de Productos ---
@login_required
@permission_required('catalogo.view_producto', raise_exception=True)
def producto_list(request):
    nombre_filtro = request.GET.get('nombre', '')
    productos = Producto.objects.all().select_related('Categorias')
    if nombre_filtro:
        productos = productos.filter(nombre__icontains=nombre_filtro)

    from django.core.paginator import Paginator
    paginator = Paginator(productos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'catalogo/producto_list.html', {'page_obj': page_obj, 'nombre_filtro': nombre_filtro})

@login_required
@permission_required('catalogo.view_producto', raise_exception=True)
def producto_detail(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'catalogo/producto_detail.html', {'producto': producto})

@login_required
@permission_required('catalogo.add_producto', raise_exception=True)
def producto_create(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.creado = timezone.now()
            producto.save()
            messages.success(request, 'Producto creado exitosamente.')
            return redirect('catalogo:producto_list')
    else:
        form = ProductoForm()
    return render(request, 'catalogo/producto_form.html', {'form': form})

@login_required
@permission_required('catalogo.change_producto', raise_exception=True)
def producto_update(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            producto_actualizado = form.save(commit=False)
            producto_actualizado.modificado = timezone.now()
            producto_actualizado.save()
            messages.success(request, 'Producto actualizado exitosamente.')
            return redirect('catalogo:producto_list')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'catalogo/producto_form.html', {'form': form})

@login_required
@permission_required('catalogo.delete_producto', raise_exception=True)
def producto_delete(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        nombre_prod = producto.nombre
        producto.delete()
        messages.success(request, f'Producto {nombre_prod} eliminado.')
        return redirect('catalogo:producto_list')
    return render(request, 'catalogo/producto_confirm_delete.html', {'object': producto})


# ----------------------------------------
# EXPORTACIÓN A EXCEL
# ----------------------------------------
@login_required
@permission_required('catalogo.view_producto', raise_exception=True)
def producto_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = [
        "ID", "Nombre", "Descripción", "Marca", "Precio", 
        "Caducidad", "Elaboración", "Tipo", "Categoría", 
        "Stock Actual", "Stock Mínimo", "Stock Máximo", 
        "Presentación", "Formato"
    ]
    ws.append(headers)

    productos = Producto.objects.all().select_related('Categorias')
    for p in productos:
        ws.append([
            p.id,
            p.nombre,
            p.descripcion or "",
            p.marca or "",
            p.precio or 0,
            p.caducidad.strftime("%Y-%m-%d") if p.caducidad else "",
            p.elaboracion.strftime("%Y-%m-%d") if p.elaboracion else "",
            p.tipo,
            p.Categorias.nombre if p.Categorias else "",
            p.stock_actual or 0,
            p.stock_minimo or 0,
            p.stock_maximo or 0,
            p.presentacion or "",
            p.formato or "",
        ])

    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('catalogo.view_categoria', raise_exception=True)
def categoria_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías"

    headers = ["ID", "Nombre", "Descripción"]
    ws.append(headers)

    categorias = Categoria.objects.all()
    for c in categorias:
        ws.append([
            c.id,
            c.nombre,
            c.descripcion or "",
        ])

    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'
    wb.save(response)
    return response
