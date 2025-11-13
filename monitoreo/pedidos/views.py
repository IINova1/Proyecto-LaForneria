from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator

# --- Importaciones ---
from .models import Pedido, DetallePedido, Cliente
from .forms import ClienteForm
from catalogo.models import Producto
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


# --------------------
# Vistas de la Tienda (Públicas)
# --------------------

def ver_productos(request):
    """
    Vista para que los clientes vean los productos disponibles.
    Incluye filtros persistentes y paginación.
    """
    DEFAULT_SORT = 'alpha_asc'
    DEFAULT_PER_PAGE = 9
    PER_PAGE_OPTIONS = [5, 9, 15, 30]

    SESSION_KEY_Q = 'productos_q'
    SESSION_KEY_SORT = 'productos_sort'
    SESSION_KEY_PER_PAGE = 'productos_per_page'

    # --- Búsqueda persistente ---
    if 'q' in request.GET:
        search_query = request.GET.get('q', '')
        request.session[SESSION_KEY_Q] = search_query
    else:
        search_query = request.session.get(SESSION_KEY_Q, '')

    # --- Orden persistente ---
    if 'sort' in request.GET:
        sort_by = request.GET.get('sort', DEFAULT_SORT)
        request.session[SESSION_KEY_SORT] = sort_by
    else:
        sort_by = request.session.get(SESSION_KEY_SORT, DEFAULT_SORT)

    # --- Paginación persistente ---
    if 'per_page' in request.GET:
        try:
            per_page = int(request.GET.get('per_page', DEFAULT_PER_PAGE))
            if per_page not in PER_PAGE_OPTIONS:
                per_page = DEFAULT_PER_PAGE
        except ValueError:
            per_page = DEFAULT_PER_PAGE
        request.session[SESSION_KEY_PER_PAGE] = per_page
    else:
        per_page = request.session.get(SESSION_KEY_PER_PAGE, DEFAULT_PER_PAGE)

    page_number = request.GET.get('page', 1)

    # --- Filtro base ---
    productos_list = Producto.objects.filter(stock_actual__gt=0)

    # --- Filtro de búsqueda ---
    if search_query:
        productos_list = productos_list.filter(
            Q(nombre__icontains=search_query) |
            Q(descripcion__icontains=search_query)
        )

    # --- Ordenamiento ---
    if sort_by == 'precio_asc':
        productos_list = productos_list.order_by('precio')
    elif sort_by == 'precio_desc':
        productos_list = productos_list.order_by('-precio')
    elif sort_by == 'alpha_desc':
        productos_list = productos_list.order_by('-nombre')
    else:
        productos_list = productos_list.order_by('nombre')

    paginator = Paginator(productos_list, per_page)
    try:
        page_obj = paginator.get_page(page_number)
    except Exception:
        page_obj = paginator.get_page(1)

    context = {
        'page_obj': page_obj,
        'paginator': paginator,
        'current_q': search_query,
        'current_sort': sort_by,
        'current_per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
    }
    return render(request, 'pedidos/ver_productos.html', context)


def agregar_al_carrito(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    try:
        cantidad = int(request.POST.get('cantidad', 1))
    except (TypeError, ValueError):
        cantidad = 1

    carrito = request.session.get('carrito', {})
    carrito[str(pk)] = carrito.get(str(pk), 0) + cantidad
    request.session['carrito'] = carrito

    messages.success(request, f'¡Producto "{producto.nombre}" agregado al carrito!')
    return redirect('pedidos:ver_carrito')


def ver_carrito(request):
    carrito = request.session.get('carrito', {})
    items_carrito = []
    total_carrito = 0

    for producto_id, cantidad in carrito.items():
        producto = get_object_or_404(Producto, pk=producto_id)
        subtotal = producto.precio * cantidad
        items_carrito.append({'producto': producto, 'cantidad': cantidad, 'subtotal': subtotal})
        total_carrito += subtotal

    return render(request, 'pedidos/ver_carrito.html', {
        'items_carrito': items_carrito,
        'total_carrito': total_carrito
    })


@login_required
def realizar_pedido(request):
    carrito = request.session.get('carrito', {})
    if not carrito:
        messages.warning(request, "Tu carrito está vacío.")
        return redirect('pedidos:ver_productos')

    total_pedido = 0
    items_para_pedido = []
    for producto_id, cantidad in carrito.items():
        producto = get_object_or_404(Producto, pk=producto_id)
        total_pedido += producto.precio * cantidad
        items_para_pedido.append((producto, cantidad))

    pedido = Pedido.objects.create(usuario=request.user, total=total_pedido)
    for producto, cantidad in items_para_pedido:
        DetallePedido.objects.create(
            pedido=pedido, producto=producto, cantidad=cantidad, precio=producto.precio
        )
        producto.stock_actual -= cantidad
        producto.save()

    request.session['carrito'] = {}
    messages.success(request, "¡Pedido realizado con éxito!")
    return redirect('pedidos:pedido_exitoso')


def pedido_exitoso(request):
    return render(request, 'pedidos/pedido_exitoso.html')


# -------------------------------
# CRUD de Clientes (solo Admins)
# -------------------------------

@login_required
def cliente_list(request):
    if not request.user.is_staff:
        return redirect('pedidos:ver_productos')

    clientes = Cliente.objects.all().order_by('idclientes')
    return render(request, 'pedidos/cliente_list.html', {'clientes': clientes})


@login_required
def cliente_create(request):
    if not request.user.is_staff:
        return redirect('pedidos:ver_productos')

    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente creado exitosamente.")
            return redirect('pedidos:cliente_list')
        else:
            messages.error(request, "Hubo errores al crear el cliente. Revisa los campos.")
    else:
        form = ClienteForm()

    return render(request, 'pedidos/cliente_form.html', {'form': form})


@login_required
def cliente_update(request, pk):
    if not request.user.is_staff:
        return redirect('pedidos:ver_productos')

    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente actualizado exitosamente.")
            return redirect('pedidos:cliente_list')
        else:
            messages.error(request, "Error al actualizar el cliente.")
    else:
        form = ClienteForm(instance=cliente)

    return render(request, 'pedidos/cliente_form.html', {'form': form})


@login_required
def cliente_delete(request, pk):
    if not request.user.is_staff:
        return redirect('pedidos:ver_productos')

    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, "Cliente eliminado exitosamente.")
        return redirect('pedidos:cliente_list')

    return render(request, 'pedidos/cliente_confirm_delete.html', {'object': cliente})


# -------------------------------
# CRUD de Pedidos (solo Admins)
# -------------------------------

@login_required
def pedido_list(request):
    if not request.user.is_staff:
        return redirect('pedidos:ver_productos')

    pedidos = Pedido.objects.all().order_by('-fecha_pedido')
    return render(request, 'pedidos/pedido_list.html', {'pedidos': pedidos})


@login_required
def exportar_pedidos_excel(request):
    if not request.user.is_staff:
        return redirect('pedidos:ver_productos')

    pedidos = Pedido.objects.all().order_by('-fecha_pedido')

    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Encabezados
    columnas = ["ID", "Cliente", "Correo", "Fecha del Pedido", "Total", "Estado"]
    for col_num, titulo in enumerate(columnas, 1):
        c = ws.cell(row=1, column=col_num)
        c.value = titulo
        c.font = openpyxl.styles.Font(bold=True)

    # Filas de datos
    for row_num, pedido in enumerate(pedidos, 2):
        ws.cell(row=row_num, column=1).value = pedido.id
        ws.cell(row=row_num, column=2).value = f"{pedido.usuario.first_name} {pedido.usuario.last_name}"
        ws.cell(row=row_num, column=3).value = pedido.usuario.email
        ws.cell(row=row_num, column=4).value = pedido.fecha_pedido.strftime("%d/%m/%Y %H:%M")
        ws.cell(row=row_num, column=5).value = float(pedido.total)
        ws.cell(row=row_num, column=6).value = pedido.estado

    # Ajustar ancho automático
    for col_num in range(1, len(columnas) + 1):
        columna = get_column_letter(col_num)
        ws.column_dimensions[columna].width = 20

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos.xlsx"'
    wb.save(response)
    return response


@login_required
def pedido_detail(request, pk):
    if not request.user.is_staff:
        return redirect('pedidos:ver_productos')

    pedido = get_object_or_404(Pedido, pk=pk)
    for detalle in pedido.detalles.all():
        detalle.subtotal = detalle.cantidad * detalle.precio

    return render(request, 'pedidos/pedido_detail.html', {'pedido': pedido})

