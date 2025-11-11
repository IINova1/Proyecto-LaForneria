from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import Proveedor
from .forms import ProveedorForm


# --- Helpers de rol ---
def _rol_usuario(user):
    """Devuelve el nombre del rol o None."""
    try:
        return user.Roles.nombre if user.Roles else None
    except Exception:
        return None


def _es_admin(user):
    return user.is_authenticated and (user.is_staff or _rol_usuario(user) == "Admin")


def _es_editor(user):
    return user.is_authenticated and _rol_usuario(user) == "Editor"


def _es_lector(user):
    return user.is_authenticated and _rol_usuario(user) == "Lector"


# --- CRUD de Proveedores --- #

@login_required
def proveedor_list(request):
    """Listar proveedores (Admin, Editor o Lector)."""
    rol = _rol_usuario(request.user)
    if rol not in ["Admin", "Editor", "Lector"]:
        return redirect('pedidos:ver_productos')

    query = request.GET.get('q', '').strip()
    page = request.GET.get('page', 1)

    proveedores_qs = Proveedor.objects.all()

    if query:
        proveedores_qs = proveedores_qs.filter(
            Q(nombre_empresa__icontains=query) |
            Q(rut__icontains=query) |
            Q(email__icontains=query)
        )

    proveedores_qs = proveedores_qs.order_by('nombre_empresa')

    paginator = Paginator(proveedores_qs, 10)
    proveedores_page = paginator.get_page(page)

    context = {
        'proveedores': proveedores_page,
        'query': query,
        'rol': rol,
    }
    return render(request, 'proveedores/proveedor_list.html', context)


@login_required
def crear_proveedor(request):
    """Crear un nuevo proveedor (solo Admin o Editor)."""
    rol = _rol_usuario(request.user)
    if rol not in ["Admin", "Editor"]:
        messages.error(request, "No tienes permisos para crear proveedores.")
        return redirect('proveedores:listar_proveedores')

    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor creado exitosamente.")
            return redirect('proveedores:listar_proveedores')
        else:
            messages.error(request, "Hay errores en el formulario.")
    else:
        form = ProveedorForm()

    return render(request, 'proveedores/proveedor_form.html', {'form': form, 'accion': 'Nuevo'})


@login_required
def editar_proveedor(request, pk):
    """Editar proveedor existente (solo Admin o Editor)."""
    rol = _rol_usuario(request.user)
    if rol not in ["Admin", "Editor"]:
        messages.error(request, "No tienes permisos para editar proveedores.")
        return redirect('proveedores:listar_proveedores')

    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor actualizado correctamente.")
            return redirect('proveedores:listar_proveedores')
        else:
            messages.error(request, "Hay errores en el formulario.")
    else:
        form = ProveedorForm(instance=proveedor)

    return render(request, 'proveedores/proveedor_form.html', {'form': form, 'accion': 'Editar'})


@login_required
def eliminar_proveedor(request, pk):
    """Eliminar proveedor (solo Admin)."""
    rol = _rol_usuario(request.user)
    if rol != "Admin":
        messages.error(request, "Solo los administradores pueden eliminar proveedores.")
        return redirect('proveedores:listar_proveedores')

    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        nombre = proveedor.nombre_empresa
        proveedor.delete()
        messages.success(request, f'Proveedor "{nombre}" eliminado correctamente.')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True})
        return redirect('proveedores:listar_proveedores')

    return redirect('proveedores:listar_proveedores')


@login_required
def exportar_proveedores_excel(request):
    """Exportar lista de proveedores (solo Admin o Editor)."""
    rol = _rol_usuario(request.user)
    if rol not in ["Admin", "Editor"]:
        messages.error(request, "No tienes permisos para exportar proveedores.")
        return redirect('proveedores:listar_proveedores')

    proveedores = Proveedor.objects.all().order_by('nombre_empresa')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    headers = ["RUT", "Empresa", "Contacto", "Email", "Teléfono", "Dirección", "Rubro"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for p in proveedores:
        ws.append([
            p.rut,
            p.nombre_empresa,
            p.nombre_contacto or "",
            p.email,
            p.telefono or "",
            p.direccion or "",
            p.rubro or "",
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="proveedores.xlsx"'
    return response
