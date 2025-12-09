from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

# Importamos los formularios y modelos desde la app local 'usuarios'
from .forms import (
    CustomRegisterForm, UserProfileForm, DireccionForm
)
from .models import Usuario, Direccion, Rol

# --------------------
# Vistas de Autenticación
# --------------------

def register(request):
    """
    Vista para el registro de nuevos usuarios (Público).
    """
    if request.method == 'POST':
        # Se agrega request.FILES para que la imagen se guarde si el usuario sube una
        form = CustomRegisterForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'¡Bienvenido, {user.first_name}! Tu cuenta ha sido creada.')
            # Redirige al nuevo usuario a la tienda
            return redirect('pedidos:ver_productos')
    else:
        form = CustomRegisterForm()
    
    return render(request, 'usuarios/register.html', {'form': form})


@login_required
def perfil(request):
    """
    Vista para que el usuario vea y actualice sus datos personales y dirección.
    """
    usuario = request.user
    # Busca la dirección existente o usa None si no existe
    try:
        direccion = usuario.Direccion
    except Direccion.DoesNotExist:
        direccion = None

    if request.method == 'POST':
        # Se usa UserProfileForm para editar datos personales + request.FILES para avatar
        user_form = UserProfileForm(request.POST, request.FILES, instance=usuario)
        direccion_form = DireccionForm(request.POST, instance=direccion)
        
        if user_form.is_valid() and direccion_form.is_valid():
            user_form.save() # Guarda los cambios en el Usuario
            
            # Guarda la dirección
            nueva_direccion = direccion_form.save(commit=False)
            if direccion is None:
                # Si no había dirección, la crea y la asigna al usuario
                nueva_direccion.save()
                usuario.Direccion = nueva_direccion
                usuario.save()
            else:
                # Si ya existía, solo la guarda (actualiza)
                nueva_direccion.save()
            
            messages.success(request, '¡Tu perfil ha sido actualizado con éxito!')
            return redirect('usuarios:perfil')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')

    else:
        # En un GET, muestra los formularios con la data existente
        user_form = UserProfileForm(instance=usuario)
        direccion_form = DireccionForm(instance=direccion)

    context = {
        'user_form': user_form,
        'direccion_form': direccion_form
    }
    return render(request, 'usuarios/perfil.html', context)


# ----------------------------------------
# Vistas Protegidas (SOLO PARA ADMINS)
# --- CRUD de Usuarios ---
# ----------------------------------------

@login_required
@permission_required('usuarios.view_usuario', raise_exception=True)
def usuario_list(request):
    """
    Lista de usuarios con filtros avanzados (Búsqueda, Rol, Estado) y paginación.
    """
    # 1. Consulta Base Optimizada
    usuarios_qs = Usuario.objects.all().select_related('Roles', 'Direccion').order_by('id')

    # 2. Capturar parámetros de la URL
    q = request.GET.get('q', '').strip()
    rol_id = request.GET.get('rol', '')
    estado = request.GET.get('estado', '')

    # 3. Aplicar Filtros Dinámicos
    if q:
        usuarios_qs = usuarios_qs.filter(
            Q(first_name__icontains=q) | 
            Q(last_name__icontains=q) | 
            Q(email__icontains=q) | 
            Q(run__icontains=q)
        )
    
    if rol_id:
        if rol_id == 'admin': # Filtro especial para superusuarios
            usuarios_qs = usuarios_qs.filter(is_superuser=True)
        elif rol_id == 'sin_rol':
            usuarios_qs = usuarios_qs.filter(Roles__isnull=True, is_superuser=False)
        else:
            usuarios_qs = usuarios_qs.filter(Roles__id=rol_id)

    if estado:
        if estado == 'activo':
            usuarios_qs = usuarios_qs.filter(is_active=True)
        elif estado == 'inactivo':
            usuarios_qs = usuarios_qs.filter(is_active=False)

    # 4. Paginación (10 usuarios por página)
    paginator = Paginator(usuarios_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 5. Obtener roles para el dropdown de filtros
    roles = Rol.objects.all()

    context = {
        'usuarios': page_obj,
        'page_obj': page_obj,
        'roles': roles,
        # Devolvemos los filtros para mantenerlos seleccionados en el HTML
        'current_q': q,
        'current_rol': rol_id,
        'current_estado': estado
    }
    return render(request, 'usuarios/usuario_list.html', context)


@login_required
@permission_required('usuarios.add_usuario', raise_exception=True)
def usuario_create(request):
    """
    Crear usuario desde el panel de administración.
    """
    if request.method == 'POST':
        form = CustomRegisterForm(request.POST) 
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('usuarios:usuario_list')
    else:
        form = CustomRegisterForm()
        
    return render(request, 'usuarios/register.html', {'form': form}) 


@login_required
@permission_required('usuarios.change_usuario', raise_exception=True)
def usuario_update(request, pk):
    """
    Editar usuario existente.
    """
    usuario = get_object_or_404(Usuario, pk=pk)
    
    if request.method == 'POST':
        # Se usa UserProfileForm para que el admin pueda editar lo mismo que el usuario
        form = UserProfileForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado exitosamente.')
            return redirect('usuarios:usuario_list')
    else:
        form = UserProfileForm(instance=usuario)
        
    return render(request, 'usuarios/perfil_admin_edit.html', {'form': form, 'usuario_editado': usuario}) 


@login_required
@permission_required('usuarios.delete_usuario', raise_exception=True)
def usuario_delete(request, pk):
    """
    Eliminar usuario.
    """
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        email_borrado = usuario.email
        usuario.delete()
        messages.success(request, f'Usuario {email_borrado} eliminado.')
        return redirect('usuarios:usuario_list')
        
    return render(request, 'usuarios/usuario_confirm_delete.html', {'object': usuario})


# -------------------------------
# EXPORTAR USUARIOS A EXCEL (SOLO ADMIN)
# -------------------------------
@login_required
@permission_required('usuarios.view_usuario', raise_exception=True)
def exportar_usuarios_excel(request):
    """
    Genera un reporte Excel con todos los usuarios registrados.
    """
    usuarios = Usuario.objects.all().order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    columnas = ["ID", "Nombre", "Apellido", "Email", "Run", "Fono", "Rol", "Dirección"]
    for col_num, titulo in enumerate(columnas, 1):
        c = ws.cell(row=1, column=col_num)
        c.value = titulo
        c.font = openpyxl.styles.Font(bold=True)

    for row_num, usuario in enumerate(usuarios, 2):
        ws.cell(row=row_num, column=1).value = usuario.id
        ws.cell(row=row_num, column=2).value = usuario.first_name
        ws.cell(row=row_num, column=3).value = usuario.last_name
        ws.cell(row=row_num, column=4).value = usuario.email
        ws.cell(row=row_num, column=5).value = usuario.run
        ws.cell(row=row_num, column=6).value = usuario.fono
        ws.cell(row=row_num, column=7).value = usuario.Roles.nombre if usuario.Roles else "Sin Rol"
        ws.cell(row=row_num, column=8).value = f"{usuario.Direccion.calle} {usuario.Direccion.numero}" if usuario.Direccion else "Sin Dirección"

    for col_num in range(1, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
    wb.save(response)
    return response